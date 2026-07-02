from flask import Blueprint, render_template, request, session, redirect, url_for, send_file
from flask_mysqldb import MySQL
from utils import admin_required, calcular_saldo_cliente
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import MySQLdb

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')

mysql = None

def set_mysql(mysql_instance):
    global mysql
    mysql = mysql_instance

def obtener_fecha_quincena():
    """Obtiene las fechas de inicio y fin de la quincena actual"""
    hoy = datetime.now().date()
    
    if hoy.day <= 15:
        inicio = hoy.replace(day=1)
        # Última quincena: 16 al 30/31
        if hoy.month == 12:
            fin = datetime(hoy.year + 1, 1, 1).date() - timedelta(days=1)
        else:
            fin = datetime(hoy.year, hoy.month + 1, 1).date() - timedelta(days=1)
    else:
        # Primera quincena: 1 al 15
        inicio = hoy.replace(day=16)
        if hoy.month == 12:
            fin = datetime(hoy.year + 1, 1, 15).date()
        else:
            fin = datetime(hoy.year, hoy.month + 1, 15).date()
    
    return inicio, fin

@reportes_bp.route('/')
@admin_required
def index():
    """Dashboard de reportes"""
    return render_template('reportes/index.html')

@reportes_bp.route('/cuentas-por-cobrar')
@admin_required
def cuentas_por_cobrar():
    """Reporte de cuentas por cobrar - tabla interactiva"""
    try:
        cursor = mysql.connection.cursor()
        
        # Obtener todos los clientes con saldo pendiente
        cursor.execute("""
            SELECT DISTINCT c.id, c.nombre, c.cedula, c.celular
            FROM clientes c
            JOIN cuentas_cobrar cc ON c.id = cc.cliente_id
            WHERE cc.estado != 'pagada'
            ORDER BY c.nombre
        """)
        clientes = cursor.fetchall()
        
        # Para cada cliente, calcular su saldo
        clientes_con_saldo = []
        total_general = 0
        
        for cliente in clientes:
            saldo = calcular_saldo_cliente(mysql, cliente['id'])
            if saldo > 0:
                clientes_con_saldo.append({
                    'id': cliente['id'],
                    'nombre': cliente['nombre'],
                    'cedula': cliente['cedula'],
                    'celular': cliente['celular'],
                    'saldo': saldo
                })
                total_general += saldo
        
        cursor.close()
        
        return render_template('reportes/cuentas_por_cobrar.html', 
                             clientes=clientes_con_saldo,
                             total_general=total_general)
    except Exception as e:
        error = f'Error: {str(e)}'
        return render_template('reportes/cuentas_por_cobrar.html', error=error, clientes=[], total_general=0.0)

@reportes_bp.route('/cuentas-por-cobrar/pdf')
@admin_required
def cuentas_por_cobrar_pdf():
    """Exportar reporte de cuentas por cobrar a PDF"""
    try:
        cursor = mysql.connection.cursor()
        
        # Obtener datos
        cursor.execute("""
            SELECT DISTINCT c.id, c.nombre, c.cedula, c.celular
            FROM clientes c
            JOIN cuentas_cobrar cc ON c.id = cc.cliente_id
            WHERE cc.estado != 'pagada'
            ORDER BY c.nombre
        """)
        clientes = cursor.fetchall()
        
        # Preparar datos
        datos = []
        total_general = 0
        
        for cliente in clientes:
            saldo = calcular_saldo_cliente(mysql, cliente['id'])
            if saldo > 0:
                datos.append([
                    cliente['cedula'],
                    cliente['nombre'],
                    cliente['celular'],
                    f"C${saldo:,.2f}"
                ])
                total_general += saldo
        
        cursor.close()
        
        # Crear PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=20,
            alignment=1  # Centro
        )
        
        # Título
        elements.append(Paragraph("REPORTE DE CUENTAS POR COBRAR", title_style))
        elements.append(Paragraph(f"Fecha: {datetime.now().strftime('%d/%m/%Y')}", 
                                styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabla
        encabezados = ['Cédula', 'Cliente', 'Celular', 'Saldo Pendiente']
        tabla_data = [encabezados] + datos + [
            ['', '', 'TOTAL:', f"C${total_general:,.2f}"]
        ]
        
        tabla = Table(tabla_data)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f9fafb')])
        ]))
        
        elements.append(tabla)
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"cuentas_por_cobrar_{datetime.now().strftime('%Y%m%d')}.pdf"
        )
    
    except Exception as e:
        error = f'Error: {str(e)}'
        return render_template('reportes/cuentas_por_cobrar.html', error=error, clientes=[], total_general=0.0)

@reportes_bp.route('/cuentas-por-cobrar/excel')
@admin_required
def cuentas_por_cobrar_excel():
    """Exportar reporte de cuentas por cobrar a Excel"""
    try:
        cursor = mysql.connection.cursor()
        
        # Obtener datos
        cursor.execute("""
            SELECT DISTINCT c.id, c.nombre, c.cedula, c.celular
            FROM clientes c
            JOIN cuentas_cobrar cc ON c.id = cc.cliente_id
            WHERE cc.estado != 'pagada'
            ORDER BY c.nombre
        """)
        clientes = cursor.fetchall()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cuentas por Cobrar"
        
        # Estilos
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        encabezados = ['Cédula', 'Cliente', 'Celular', 'Saldo Pendiente']
        for col, header in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        total_general = 0
        row = 2
        
        for cliente in clientes:
            saldo = calcular_saldo_cliente(mysql, cliente['id'])
            if saldo > 0:
                ws.cell(row=row, column=1).value = cliente['cedula']
                ws.cell(row=row, column=2).value = cliente['nombre']
                ws.cell(row=row, column=3).value = cliente['celular']
                ws.cell(row=row, column=4).value = saldo
                
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border
                    if col == 4:
                        ws.cell(row=row, column=col).number_format = 'C$#,##0.00'
                
                total_general += saldo
                row += 1
        
        # Total
        ws.cell(row=row, column=3).value = "TOTAL:"
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4).value = total_general
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).number_format = 'C$#,##0.00'
        
        # Ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        
        cursor.close()
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"cuentas_por_cobrar_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    
    except Exception as e:
        error = f'Error: {str(e)}'
        return render_template('reportes/cuentas_por_cobrar.html', error=error, clientes=[], total_general=0.0)

@reportes_bp.route('/facturas')
@admin_required
def facturas():
    """Reporte de facturas por período"""
    try:
        fecha_inicio = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        estado = request.args.get('estado', 'todos')
        
        cursor = mysql.connection.cursor()
        
        query = """
            SELECT f.id, f.numero_factura, f.fecha_emision, c.nombre as cliente,
                   f.total, f.estado, u.nombre as vendedor
            FROM facturas f
            JOIN clientes c ON f.cliente_id = c.id
            JOIN usuarios u ON f.vendedor_id = u.id
            WHERE 1=1
        """
        params = []
        
        if fecha_inicio:
            query += " AND f.fecha_emision >= %s"
            params.append(fecha_inicio)
        
        if fecha_fin:
            query += " AND f.fecha_emision <= %s"
            params.append(fecha_fin)
        
        if estado != 'todos':
            query += " AND f.estado = %s"
            params.append(estado)
        
        query += " ORDER BY f.fecha_emision DESC"
        
        cursor.execute(query, params)
        facturas = cursor.fetchall()
        
        cursor.close()
        
        return render_template('reportes/facturas.html', 
                             facturas=facturas,
                             fecha_inicio=fecha_inicio,
                             fecha_fin=fecha_fin,
                             estado=estado)
    except Exception as e:
        error = f'Error: {str(e)}'
        return render_template('reportes/facturas.html', error=error, facturas=[], fecha_inicio='', fecha_fin='', estado='todos')

@reportes_bp.route('/inventario')
@admin_required
def inventario():
    """Reporte de inventario con opción de alertar bajo stock"""
    try:
        bajo_stock = request.args.get('bajo_stock', 'false').lower() == 'true'
        
        cursor = mysql.connection.cursor()
        
        if bajo_stock:
            cursor.execute("""
                SELECT id, nombre, descripcion, precio_venta, stock, activo 
                FROM productos 
                WHERE stock <= 10
                ORDER BY stock ASC, nombre
            """)
        else:
            cursor.execute("""
                SELECT id, nombre, descripcion, precio_venta, stock, activo 
                FROM productos 
                ORDER BY stock ASC, nombre
            """)
            
        productos = cursor.fetchall()
        cursor.close()
        
        return render_template('reportes/inventario.html', productos=productos, bajo_stock=bajo_stock)
        
    except Exception as e:
        error = f'Error: {str(e)}'
        return render_template('reportes/inventario.html', error=error, productos=[], bajo_stock=False)
